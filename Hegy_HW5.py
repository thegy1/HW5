# Tyler Hegy
# UWYO COSC 1010
# Submission Date: 11/17/2024
# HW 05
# Lab Section: 12
# Sources, people worked with, help given to: N/A
# Comments: None

import openpyxl
from openpyxl.styles import Color, PatternFill
import string
wb= openpyxl.Workbook()
sheet = wb.active

red = Color(rgb='ff0000')
fill_r= PatternFill(patternType='solid',fgColor=red)

blue = Color(rgb='0000ff')
fill_b= PatternFill(patternType='solid',fgColor=blue)

black = Color(rgb='000000')
fill_k= PatternFill(patternType='solid',fgColor=black)

gray = Color(rgb='808080')
fill_gray= PatternFill(patternType='solid',fgColor=gray)

white = Color(rgb='ffffff')
fill_w= PatternFill(patternType='solid',fgColor=white)

green = Color(rgb='008000')
fill_g= PatternFill(patternType='solid',fgColor=green)

orange = Color(rgb = 'ffa500')
fill_o= PatternFill(patternType='solid',fgColor=orange)

cells_k=['A1','A2','A3','A4','A5','A6','A7','A8','A9','A10', 
         'A11','A12','A13','B1','B13', 'C1','C13', 'D1','D13', 'E1','E13', 'F1','F13',
         'G1','G13', 'H1','H13', 'I1','I13','J1','J2','J3','J4','J5',
         'J6','J7','J8','J9','J10','J11','J12','J13']
cells_r=['F2','F3','F4','G2']
cells_b = ['D5','D6','C6','E6','I4','I5','I6','I7']
cells_o = ['B6','B7','C7','D7']
cells_g = ['G6','G7','H7','H6']
cells_gray = ['B8','C8','D8','E8','F8','G8','H8','I8',
              'B12','C12','D12','E12','F12','G12','H12','I12',
              'B9','D9','E9','F9','G9','H9','I9',
              'B11','D11','E11','F11','G11','H11','I11',
              'E10','E10']

for chr in string.ascii_uppercase[:10]:
    sheet.column_dimensions[chr].width = 4
    for i in range(1,21):
        sheet.row_dimensions[i].height= 24
        coord = chr+str(i)
        if coord in cells_k:
            sheet[coord].fill = fill_k
        elif coord in cells_r:
            sheet[coord].fill=fill_r
        elif coord in cells_b:
            sheet[coord].fill=fill_b
        elif coord in cells_o:
            sheet[coord].fill=fill_o
        elif coord in cells_g:
            sheet[coord].fill=fill_g
        elif coord in cells_gray:
            sheet[coord].fill=fill_gray
        else:
            sheet[coord].fill=fill_w
#Tetris on GameBoy
wb.save('pixelart.xlsx')
